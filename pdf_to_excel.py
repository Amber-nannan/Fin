# -*- coding: UTF-8 –*-
from openpyxl import load_workbook
from pathlib import Path
import pdfplumber
import re

# 根据表格模板提取pdf数据，保存到每个文件夹下
def get_tables(file):
    pdf = pdfplumber.open(str(file))
    tables = []
    for page in pdf.pages:
        table = page.extract_tables()

        if not table:
            continue
        for row in table:
            tables += row
    pdf.close()
    tables_new = [
        [str(cell).replace('\n', '') for cell in row] for row in tables]

    return tables_new


def get_code_name(file: Path):
    if isinstance(file, str):
        file = Path(file)
    name = file.stem.split('_')[0]
    if name.startswith('1'):
        name = name + '.SZ'
    elif name.startswith('5'):
        name = name + '.SH'
    return name


def switch_data_format(values):
    float_values = []
    for v in values:
        if isinstance(v, (float, int)):
            float_values.append(v)
        elif v != '':
            try:
                if v.endswith('%'):
                    f = float(v[:-1]) / 100
                else:
                    f = float(v.replace(',', '').replace('\n', ''))
                float_values.append(f)
            except ValueError:
                float_values.append(v)
        else:
            float_values.append('')

    return float_values




def check_string_contains(my_string, substrings_to_check):
    # my_string = "本期收入为其他收入"
    # substrings_to_check = ['收入', '其他收入', '发电收入', '租金收入']
    pattern = '|'.join(substrings_to_check)
    return bool(re.search(pattern, my_string))


def get_data(file):
    tables = get_tables(file)
    A = B = C = D = E = F = G = H = I = J = K = L = M = N = O = P = Q = R = S = T = U = V = W = X = ''
    A = get_code_name(file)
    G_, H_ = [], []
    for n, row in enumerate(tables):

        if row[0] == "基金简称":
            B = row[-1]
        if '主要财务指标' in row:
            for row2 in tables[n:n + 10]:
                if isinstance(row2[0], str) and '本期收入' in row2[0]:
                    for r in row2:
                        if isinstance(r, str) and ',' in r:
                            D = r
                if isinstance(row2[0], str) and '本期净利润' in row2[0]:
                    for r in row2:
                        if isinstance(r, str) and ',' in r:
                            E = r
                if isinstance(row2[0], str) and '本期经营活动' in row2[0].replace('\n', ''):
                    for r in row2:
                        if isinstance(r, str) and ',' in r:
                            F = r
        # 最原始的代码
        # for rr in row:
        #     if isinstance(rr,str) and ('其他收入' in rr.replace('\n','') or '租金收入' in rr.replace('\n','')):
        #         for row2 in tables[n:n+4]:
        #             if '合计' in row2:
        #                 for r in row2:
        #                     if isinstance(r,str) and ',' in r:
        #                         G = r
        #                         break
        for rr in row:
            words_to_check = ['其他收入', '租金收入', '收入']
            if isinstance(rr, str) and check_string_contains(rr, words_to_check):
                # print(rr,"yes")
                for row2 in tables[n:n + 5]:
                    if '合计' in row2:
                        # print(row2,'yes')
                        for r in row2:
                            if isinstance(r, str) and check_string_contains(r, ['100']):
                                for rn in row2:
                                    if isinstance(rn, str) and ',' in rn:
                                        G0 = float(rn.replace(',', ''))
                                        G_.append(G0)
                                        G_ = list(set(G_))
                                        break

        for rr in row:
            if isinstance(rr, str) and '其他成本' in rr.replace('\n', ''):
                for row2 in tables[n:n + 4]:
                    if '合计' in row2:
                        for r in row2:
                            if isinstance(r, str) and ',' in r:
                                H0 = float(r.replace(',', ''))
                                H_.append(H0)
                                break

        if '可供分配金额' in row and '单位可供分配金额' in row:
            for row2 in tables[n:n + 4]:
                if '本期' in row2:
                    for r in row2:
                        if ',' in r:
                            try:
                                I = float(r.replace(',', ''))
                            except:
                                pass
                        try:
                            K = float(r)
                        except Exception:
                            pass
                if '本年累计' in row2:
                    for r in row2:
                        if ',' in r:
                            try:
                                J = float(r.replace(',', ''))

                            except:
                                pass
                        try:
                            L = float(r)
                        except Exception:
                            pass

        if row[0] and '折旧和摊销' in row[0].replace('\n', ''):
            for r in row[1:]:
                if r and ',' in r:
                    M = r
            for r in tables[n + 1]:
                if r and ',' in r:
                    N = r
            for r in tables[n + 2]:
                if r and ',' in r:
                    O = r
            for r in tables[n + 3]:
                if r and ',' in r:
                    P = r
        if row[0] == "调增项":
            Q = 0
            for row3 in tables[n:]:
                for r in row3:
                    if isinstance(r, str) and ',' in r:
                        try:
                            q = float(r.replace(',', ''))
                        except Exception:
                            q = 0
                        Q += q
                        continue
                if row3[0] == '调减项':
                    break

        if row[0] == "调减项":
            R = 0
            ss = 0
            for row3 in tables[n:]:
                if ss == 1:
                    break
                for r in row3:
                    if isinstance(r, str) and '分配金额' in r:
                        ss = 1
                        break
                    if isinstance(r, str) and ',' in r:
                        try:
                            q = float(r.replace(',', ''))
                        except Exception:
                            q = 0
                        R += q
                        continue

        for rr in row:
            check_list = ['期初基金份额']
            if check_string_contains(rr, check_list):
                S = row[-1]
                if check_string_contains(rr, ['份']):
                    S = S.replace('份', '')


        for rr in row:
            check_list = ['期末基金份额']
            if check_string_contains(rr, check_list):
                T = row[-1]
                if check_string_contains(rr, ['份']):
                    T = T.replace('份', '')


        for rr in row:
            check_list = ['期初管理人持有的本基金份额']
            if check_string_contains(rr, check_list):
                U = row[-1]
        for rr in row:
            check_list = ['期初管理人持有的本基金份额']
            if check_string_contains(rr, check_list):
                V = row[-1]
        for rr in row:
            check_list1 = ['本基金份额占基金总份额比例']
            check_list2 = ['%']
            if check_string_contains(rr, check_list1) and check_string_contains(rr, check_list2):
                W = row[-1]
        for rr in row:
            check_list1 = ['银行存款和结算备付金']
            if check_string_contains(rr, check_list1):
                row = [x for x in row if x not in ['None','']]
                X = row[-1]

    G = sum(G_)
    H = sum(H_)

    # 去掉逗号与换行
    values = [A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X]
    num_values = switch_data_format(values)

    return num_values


def run(root_dir, base_dir):
    wb = load_workbook(root_dir + '/models1.xlsx')
    sheet = wb.active
    row = 3
    for file in Path(base_dir).rglob('*.pdf'):  # 做了一下尝试修改!!!原来为base_dir
        lst = get_data(file)
        # print(lst)
        # print(file,'提取完毕')
        for n, r in enumerate(lst):
            sheet.cell(row=row, column=n + 1).value = r
        row += 1
    name = base_dir.joinpath(base_dir.stem + '.xlsx')
    wb.save(name)


def main(root_dir, update):
    for base_dir in Path(root_dir).glob('*'):
        if base_dir.is_dir():
            name = base_dir.joinpath(base_dir.stem + '.xlsx')
            if update == 'part':
                if name.exists():
                    print(name, '已存在')
                else:
                    run(root_dir, base_dir)
                    print(name, '保存成功。')
            else:
                run(root_dir, base_dir)
                print(name, '保存成功。')

# if __name__ == '__main__':
#     root_dir ='PDF'
#     main(root_dir)  
# file = 'Fin_rp/Qreport_PDF/2022Q4/Q_report/508027_2022Q4.pdf'
# t = get_tables(file)
# for i in t:
#     print(i)
