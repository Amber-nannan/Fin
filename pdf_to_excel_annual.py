# -*- coding: UTF-8 –*-
from openpyxl import load_workbook
from pathlib import Path
import pdfplumber
import re

# 根据目录缩小extract页码范围: §2--5.1、13.4-§14
import pdfplumber
def get_tables(file):
    pdf = pdfplumber.open(str(file))
    toc_pages = pdf.pages[2:5]  # 这几页是目录页
    toc_page_lines=[]
    for toc_page in toc_pages:
        toc_page_lines += toc_page.extract_text_lines()
    for line in toc_page_lines:
        text=line['text']
        if '基金简介' in text:
            start_page=int(re.findall('\.{2,}\s?(\d+)',text)[0])
        if '5.2' in text  and '债券投资组合' in text:
            end_page=int(re.findall('(\d{2})',text)[0])
        if ('13.1' in text) or ('户数及持有人结构' in text):    # 这两个基本在同一页，可能都有/没有，或可能有一个
            try:
                page_=int(re.findall('\.{2,}\s?(\d+)',text)[0])
                break
            except:
                pass
    # 提取§3小节-5.1的表格 和 3.4的管理费文本
    tables = []
    content=''
    for page in pdf.pages[start_page-1:end_page]:  # 第一页是page[0]所以start_page-1
        table = page.extract_tables()
        text = page.extract_text()
        content += text
        if not table:
            continue
        for row in table:
            tables += row
    tables_new = [
        [str(cell).replace('\n', '') for cell in row] for row in tables]
    
    lines = content.splitlines()
    for line in lines:
        if '3.4' in line and '基金及资产支持证券费用收取' in line:
            a = lines.index(line)
        if '4' in line and '基础设施项目运营情况' in line:
            b = lines.index(line)
            break
    content = ''.join(lines[a+1:b])

    # 提取基金份额变动情况附近页，即13.4-§14
    tables_2= []
    if page_:
        for page in pdf.pages[page_-2:]:
            table = page.extract_tables()
            if not table:
                continue
            for row in table:
                tables_2 += row
    tables_2_new = [
        [str(cell).replace('\n', '') for cell in row] for row in tables_2]
    pdf.close()
    return tables_new, content,tables_2_new

def get_code_name(file: Path):
    if isinstance(file, str):
        file = Path(file)
    name = file.stem.split('_')[0]
    if name.startswith('1'):
        name +='.SZ'
    elif name.startswith('5'):
        name +='.SH'
    return name

def switch_data_format(values):
    float_values = []
    for v in values:
        if isinstance(v, (float, int)):
            float_values.append(v)
        elif v != '':
            try:
                if v.endswith('%'):
                    f = float(v[:-1]) / 100
                else:
                    f = float(v.replace(',', '').replace('\n', ''))
                float_values.append(f)
            except ValueError:
                float_values.append(v)
        else:
            float_values.append('')
    return float_values

def check_string_contains(my_string, substrings_to_check):
    # my_string = "本期收入为其他收入"
    # substrings_to_check = ['收入', '其他收入', '发电收入', '租金收入']
    pattern = '|'.join(substrings_to_check)
    return bool(re.search(pattern, my_string))


def get_data(file):
    tables,mgmt_fee_text,tables_2 = get_tables(file)
    A = B = C = D = E = F = G = H = I = J = K = L = M = N = O = P =\
    Q = R = S = T = U = V = W = X = Y = Z = AA = AB = AC = AD = AE = ''
    A = get_code_name(file) 
    J_, K_ = [], []
    for n, row in enumerate(tables):
        if row[0] == "基金简称":
            B = row[1]    
        if '期间数据和指标' in row:  
            for row2 in tables[n:n + 7]:
                if '本期收入' in row2[0]:
                    for r in row2:
                        if ',' in r:
                            D = r
                            continue
                if '本期净利润' in row2[0]:
                    for r in row2:
                        if ',' in r:
                            E = r
                            continue
                if '本期经营活动' in row2[0]:  
                    for r in row2:
                        if ',' in r:
                            F = r

        if '期末数据和指标' in row:  
            for row2 in tables[n:n + 5]:
                if '期末基金总资产' in row2[0]:
                    for r in row2:
                        if ',' in r:
                            G = r
                            continue
                if '期末基金净资产' in row2[0]:
                    for r in row2:
                        if ',' in r:
                            H = r
                            continue
                if '总资产与净资产的比例' in row2[0]: 
                    for r in row2:
                        if '.' in r:
                            I = r
                            if '%' not in I:
                                I += '%'
                            continue
        
        # J 为营业收入
        for rr in row:
            words_to_check = ['其他收入', '租金收入', '收入']
            if isinstance(rr, str) and check_string_contains(rr, words_to_check):
                for row2 in tables[n:n + 5]:
                    if '合计' in row2:
                        for r in row2:
                            if isinstance(r, str) and check_string_contains(r, ['100']):
                                for rn in row2:
                                    if isinstance(rn, str) and ',' in rn:
                                        J0 = float(rn.replace(',', ''))
                                        J_.append(J0)
                                        J_ = list(set(J_))
                                        break
        # K 为营业成本
        for rr in row:
            if isinstance(rr, str) and '其他成本' in rr.replace('\n', ''):
                for row2 in tables[n:n + 4]:
                    if '合计' in row2:
                        for r in row2:
                            if isinstance(r, str) and ',' in r:
                                K0 = float(r.replace(',', ''))
                                K_.append(K0)
                                break

        if '数据和指标' in row:
            
            for row2 in tables[n:n+12]:
                if '期末基金份额净值' in row2[0]:
                    for r in row2:
                        if '.' in r:
                            L = r
                            continue
                if '公允价值参考净值' in row2[0]:
                    for r in row2:
                        if '.' in r:
                            M = r
                            continue
        
        if '可供分配金额' in row and '单位可供分配金额' in row:
            for row2 in tables[n:n + 3]: 
                if '本期' in row2:
                    for r in row2:
                        if ',' in r and not re.search('[\u4e00-\u9fff]', r):
                            # break
                            N = r
                            continue
                        if '.' in r and not re.search('[\u4e00-\u9fff]', r):
                            O = r
        if '实际分配金额' in row and '单位实际分配金额' in row:
            for row2 in tables[n:n + 3]:
                if '本期' in row2:
                    for r in row2:
                        if ',' in r and '.' in r and not re.search('[\u4e00-\u9fff]', r):  # 这里可能有缺陷，有的P和L可能是0.00这样的数
                            P = r
                            continue
                        if '.' in r and not re.search('[\u4e00-\u9fff]', r):
                            Q = r
                            if float(Q)==0:
                                P='0.00'
    
        if '折旧和摊销' in row[0]:
            for r in row[1:]:
                if r and ',' in r:
                    R = r
            for r in tables[n + 1]:
                if r and ',' in r:
                    S = r
            for r in tables[n + 2]:
                if r and ',' in r:
                    T = r
            for r in tables[n + 3]:
                if r and ',' in r:
                    U = r

        if row[0] == "调增项":
            V = 0
            for row3 in tables[n:]:
                for r in row3:
                    if ',' in r:
                        try:
                            v = float(r.replace(',', ''))
                        except Exception:
                            v = 0
                        V += v
                        continue
                if row3[0] == '调减项':
                    break

        if row[0] == "调减项":  # 看起来可以简化,有空弄
            W = 0
            ss = 0
            for row3 in tables[n:]:
                if ss == 1:
                    break
                for r in row3:
                    if '分配金额' in r:
                        ss = 1
                        break
                    if ',' in r:
                        try:
                            q = float(r.replace(',', ''))
                        except Exception:
                            q = 0
                        W += q
                        continue

        if '货币资金和结算备付金合计' in row:
            for r in row:
                if r and ',' in r:
                    X = r
        
        for rr in row:
            check_list = ['期末基金份额总额']
            if check_string_contains(rr, check_list):
                AC = row[-1]
                if check_string_contains(rr, ['份']):
                    AC = AC.replace('份', '')
        if tables_2:
            for n, row in enumerate(tables_2):
                if '机构投资者' in row and '个人投资者' in row:
                    # break
                    original_list = tables_2[n+1:n+10]
                    for row2 in original_list:
                        if '.' in str(row2) and not re.search('[\u4e00-\u9fff]', str(row2)):
                            fil = [item for item in row2 if ',' in item or '.' in item]
                            Y,Z,AA,AB = fil[0],fil[1],fil[2],fil[4]
                
                if '基金管理人所有从业人员持有本基金' in row:
                    fil = [item for item in row if '.' in item]
                    AD = fil[0]
                    AE = fil[1]
                # if '基金合同生效日' in row[0] and '基金份额总额' in row[0]:
                #     AA=row[1]
                # if '期初基金份额总额' in row[0]:
                #     AB=row[1]
                if '期末基金份额总额' in row[0]:
                    AC=row[1]

        # for rr in row:
        #     check_list1 = ['本基金份额占基金总份额比例']
        #     check_list2 = ['%']
        #     if check_string_contains(rr, check_list1) and check_string_contains(rr, check_list2):
        #         W = row[-1]
                    
    J = sum(J_)
    K = sum(K_)
    # # 去掉逗号与换行
    values = [A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U, V, W, X, Y, Z, AA, AB, AC, AD ,AE]
    num_values = switch_data_format(values)
    return num_values

def run(base_dir):
    wb = load_workbook('modelsA.xlsx')
    sheet = wb.active
    row = 3
    for file in Path(base_dir).rglob('*.pdf'): 
        lst = get_data(file)
        # print(lst)
        # print(file,'提取完毕')
        for n, r in enumerate(lst):
            sheet.cell(row=row, column=n + 1).value = r
        row += 1
    name = base_dir.joinpath(base_dir.stem + '.xlsx')   #保存在文件很少的那一层文件夹
    wb.save(name)

# 该函数检查/更新所有年度的提取
def main(root_dir, update):
    # 这个root_dir='Areport_PDF'
    for base_dir in Path(root_dir).glob('*'):
        if base_dir.is_dir():
            name = base_dir.joinpath(base_dir.stem + '.xlsx')
            if update == 'part':
                if name.exists():
                    print(name, '已存在')
                else:
                    run(base_dir)
                    print(name, '保存成功。')
            else:
                run(base_dir)
                print(name, '保存成功。')

if __name__ == '__main__':
    main('Areport_PDF',1)
    # get_data(r'F:\REITs\to_intern_0903\Fin_rp\Areport_PDF\2022A\A_report\508056_2022A.pdf')

